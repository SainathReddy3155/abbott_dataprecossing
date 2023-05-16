# -*- coding: utf-8 -*-
"""
Created on Tue May 16 10:58:05 2023

@author: anuro
"""

import pandas as pd
import random
from openpyxl import load_workbook
import re
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

excel_file=pd.read_excel(r"C:\Users\anuro\OneDrive\Desktop\abbott_preprocessing1.xlsx")

mul="MUL_"
client="ABB_"


#start generating random id
def generate_unique_id(x):
    
    caps_alphabets="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    alphabets="abcdefghijklmnopqrstuvwxyz"
    numbers='0123456789'
    length=10
    generating_id= ''.join((random.choice(caps_alphabets+alphabets+numbers)) for x in range(length))
    unique_id=str(mul)+str(client)+str(generating_id)
    print(unique_id)
    return unique_id
    
##end of generating random id 


##start for proper casing and numbers function

def propercasing_number_removing(text):
    excel_file["Account: Account Name"]=excel_file["Account: Account Name"].str.replace(r"\d+","")
    excel_file["Account: Account Name"]=excel_file["Account: Account Name"].str.title()
    print(excel_file["Account: Account Name"])
    number_removed=text.str.replace("r\d+",'')
    proper_cased=number_removed.str.title()
    
    
    return proper_cased

#end of proper casing and numbers function

def removing_dr():
    text="dr. mrind1a johari"
    ptrn = re.compile(fr"^\s*(?:M(?:iss|rs?|s)|Dr|Rev|Er)\b[\s.]*", flags=re.I)
    removed_dr = [re.sub(ptrn,"",text.lower())]
    removed_dr = removed_dr[0]
    print(removed_dr)
    
    return removed_dr

##adding Dr infront of name function starts

def adding_dr():
    text="Sainath Reddy"
    string="Dr. "
    final_text=string+str(text)
    print(final_text)
    return final_text

##adding Dr infront of name function ends

def specilaity_format(speciality):
    speciality="MBBS MD DNB-chest"
    spliting_speciality=speciality.split(" ")
    # print(spliting_speciality)
    final_speciality=",".join(spliting_speciality)
    print(final_speciality)
    return final_speciality

#filling #NA,na,0 into not avialable
excel_file=excel_file.fillna("Not avialable")
print(excel_file["Account: MCI"])
#end of filling na with not available


def tally_blanks():
    file_path=r"C:\Users\anuro\OneDrive\Desktop\abbott_preprocessing1.xlsx"
    sheet_name = 'Empty counts'


    counts = excel_file.apply(lambda x: (x == 'Not available').sum())

    # Load the workbook
    book = load_workbook(file_path)
    
    # Select the sheet by name
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    
    # Write the counts to the same sheet
    counts_df = pd.DataFrame({'Column': counts.index, 'Count': counts.values})
    counts_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=writer.sheets[sheet_name].max_row)
    
    # Save the changes
    writer.save()
    writer.close()
    
    print("Counts saved to the same Excel sheet.")


def pincodes():
    pincode=50007
    match = re.search(r'\d{6}', str(pincode))
    number=""
    if match:
        number = match.group(0)
        print(number)
    else:
        number = "not avialable"
        print(number)
    return number

def email_checking():
    username="email"
    text="sainath11@gmail.com"
    match = re.match(r"(.*?)@", text)
    extracted_text=""
    if match:
        extracted_text = match.group(1)
    else:
        extracted_text="Not available"
        
    # print(type(extracted_text))
    extracted_text=re.sub(r'\d+', '', extracted_text)
    print(extracted_text)
    if extracted_text!="Not available":
        ratio=fuzz.ratio(username,extracted_text)
        match_ratio=str(ratio)+" %"
        print("match ratio : ",match_ratio)
    else:
        match_ratio="0 %"
        print(match_ratio)
        
def phonenumber_validation():
    text = "17995683155"
    if len(text)==12:
        # Remove "91" if it is at the beginning of the string
        result = re.sub(r'^91', '', text)
    elif len(text)==11:
        result = re.sub(r'^0', '', text) 
    else:
        result=text
    print(result)

def name_status():
    name=" r sainath reddy"
    
    spliting=name.split(" ")
    status=""
    flag=False
    for i in spliting:
        if len(i)==1 or len(i)==2:
            flag=True
    if flag==True:
        status="Partial Name"
        print(status)
        
    else:
        status="Full Name"
        print(status)
    return status



df = pd.DataFrame({'Specialty by Qualification': ["dentist", "Cardiologist", "Dentist","Dentist","Dentist","Dentist","Dentist","Dentist"],
                   'Qualification1': ['MBBS', 'MBBS', 'MD','BDS','MDS','DNB','FRCH','MBBS,MD']})
df
for index, row in df.iterrows():
    if row['Specialty by Qualification'].lower() == 'dentist' and row['Qualification1'].lower() not in ['mds', 'bds']:
        print("Invalid qualification")
    else:
        print("No error")
    

        
        





        
    
    
        
        




    

        

    




            
        
    
    





    
